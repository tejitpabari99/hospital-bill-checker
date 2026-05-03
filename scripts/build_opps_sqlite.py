#!/usr/bin/env python3
"""
Build data/opps.sqlite from CMS OPPS Addendum A and B files.

Sources (April 2026):
  Addendum B: https://www.cms.gov/files/zip/cy-2026-april-opps-addendum-b.zip
  Addendum A: https://www.cms.gov/files/zip/cy-2026-april-opps-addendum.zip

Usage:
  python3 scripts/build_opps_sqlite.py
  python3 scripts/build_opps_sqlite.py --addb /path/to/addendum-b.zip --adda /path/to/addendum-a.zip
"""

from __future__ import annotations

import argparse
import io
import os
import re
import sqlite3
import sys
import urllib.request
import zipfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    sys.exit(1)

DB_PATH = Path(__file__).parent.parent / "data" / "opps.sqlite"

ADDB_URL = "https://www.cms.gov/files/zip/cy-2026-april-opps-addendum-b.zip"
ADDA_URL = "https://www.cms.gov/files/zip/cy-2026-april-opps-addendum.zip"

QUARTER = "2026Q2"
EFFECTIVE_DATE = "2026-04-01"

CODE_PATTERN = re.compile(r"^(?:[0-9]{5}|[0-9]{4}[A-Z]|[A-Z][0-9]{4})$")


def download_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        return resp.read()


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS opps_addendum_b (
            quarter                         TEXT NOT NULL,
            effective_date                  TEXT NOT NULL,
            hcpcs_code                      TEXT NOT NULL,
            short_descriptor                TEXT,
            status_indicator                TEXT,
            apc                             TEXT,
            relative_weight                 NUMERIC,
            payment_rate                    NUMERIC,
            national_unadjusted_copayment   NUMERIC,
            minimum_unadjusted_copayment    NUMERIC,
            ira_coinsurance_percentage      NUMERIC,
            adjusted_beneficiary_copayment  NUMERIC,
            pass_through_expiration         TEXT,
            note                            TEXT,
            changed_flag                    TEXT,
            source_file                     TEXT,
            PRIMARY KEY (quarter, hcpcs_code)
        );

        CREATE INDEX IF NOT EXISTS idx_opps_b_code
            ON opps_addendum_b(hcpcs_code);

        CREATE INDEX IF NOT EXISTS idx_opps_b_apc
            ON opps_addendum_b(apc);

        CREATE INDEX IF NOT EXISTS idx_opps_b_si
            ON opps_addendum_b(status_indicator);

        CREATE TABLE IF NOT EXISTS opps_addendum_a (
            quarter                         TEXT NOT NULL,
            effective_date                  TEXT NOT NULL,
            apc                             TEXT NOT NULL,
            group_title                     TEXT,
            status_indicator                TEXT,
            relative_weight                 NUMERIC,
            payment_rate                    NUMERIC,
            national_unadjusted_copayment   NUMERIC,
            minimum_unadjusted_copayment    NUMERIC,
            ira_coinsurance_percentage      NUMERIC,
            adjusted_beneficiary_copayment  NUMERIC,
            pass_through_expiration         TEXT,
            note                            TEXT,
            source_file                     TEXT,
            PRIMARY KEY (quarter, apc)
        );

        CREATE INDEX IF NOT EXISTS idx_opps_a_apc
            ON opps_addendum_a(apc);
    """)


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def find_col(header_row, *candidates: str) -> int | None:
    """Find column index matching any candidate (case-insensitive, partial match allowed)."""
    for i, cell in enumerate(header_row):
        cell_upper = str(cell).strip().upper() if cell is not None else ""
        for candidate in candidates:
            if candidate.upper() in cell_upper:
                return i
    return None


def parse_addendum_b(xlsx_bytes: bytes, source_file: str) -> list[tuple]:
    """Parse Addendum B. Returns rows for opps_addendum_b."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Find the header row (contains the expected Addendum B columns)
    header_row_idx = None
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_strs = [str(c).strip().upper() if c is not None else "" for c in row]
        if (
            any(s in ("HCPCS", "HCPCS CODE") for s in row_strs)
            and any(s in ("SI", "STATUS INDICATOR", "SI / STATUS INDICATOR") for s in row_strs)
            and "PAYMENT RATE" in row_strs
        ):
            header_row_idx = i
            header_row = row
            break

    if header_row is None:
        print("ERROR: Could not find Addendum B header row")
        return []

    print(f"  Addendum B header at row {header_row_idx}: {[str(c)[:20] for c in header_row[:10]]}")

    hcpcs_col = find_col(header_row, "HCPCS")
    desc_col = find_col(header_row, "SHORT DESCRIPTOR", "SHORT DESC", "DESCRIPTOR")
    si_col = find_col(header_row, "STATUS INDICATOR", "SI")
    apc_col = find_col(header_row, "APC")
    rw_col = find_col(header_row, "RELATIVE WEIGHT", "REL WEIGHT")
    rate_col = find_col(header_row, "PAYMENT RATE", "PAYMENT", "RATE")
    cop_col = find_col(header_row, "NATIONAL UNADJ", "NATIONAL COPAY")
    min_cop_col = find_col(header_row, "MINIMUM UNADJ", "MIN COPAY")
    ira_col = find_col(header_row, "IRA", "COINSURANCE %")
    adj_cop_col = find_col(header_row, "ADJUSTED BENEF", "ADJ COPAY")
    exp_col = find_col(header_row, "PASS THROUGH EXP", "EXPIRATION")
    note_col = find_col(header_row, "NOTE")
    changed_col = find_col(header_row, "CHANGED", "FLAG")

    rows: list[tuple] = []
    data_started = False
    rows_attempted = 0

    for row in ws.iter_rows(values_only=True):
        if not data_started:
            # Detect data start: look for a cell matching a HCPCS code pattern
            # in the HCPCS column position. More robust than exact header row equality.
            if hcpcs_col is not None and hcpcs_col < len(row) and row[hcpcs_col] is not None:
                candidate = str(row[hcpcs_col]).strip().upper()
                if CODE_PATTERN.match(candidate):
                    data_started = True
                    # Fall through — process this row as data
                else:
                    continue
            else:
                continue

        if hcpcs_col is None or hcpcs_col >= len(row):
            continue

        rows_attempted += 1
        code = str(row[hcpcs_col]).strip().upper() if row[hcpcs_col] is not None else ""
        if not CODE_PATTERN.match(code):
            continue

        def get(col) -> str | float | None:
            if col is None or col >= len(row):
                return None
            return row[col]

        rows.append((
            QUARTER, EFFECTIVE_DATE, code,
            str(get(desc_col) or "").strip() or None,
            str(get(si_col) or "").strip() or None,
            str(get(apc_col) or "").strip() or None,
            safe_float(get(rw_col)),
            safe_float(get(rate_col)),
            safe_float(get(cop_col)),
            safe_float(get(min_cop_col)),
            safe_float(get(ira_col)),
            safe_float(get(adj_cop_col)),
            str(get(exp_col) or "").strip() or None,
            str(get(note_col) or "").strip() or None,
            str(get(changed_col) or "").strip() or None,
            source_file,
        ))

    wb.close()
    MIN_EXPECTED_OPPS_ROWS = 1_000
    if len(rows) < MIN_EXPECTED_OPPS_ROWS:
        raise ValueError(
            f"OPPS: only {len(rows):,} rows parsed (expected >= {MIN_EXPECTED_OPPS_ROWS:,}). "
            "Header detection may have failed — check Excel column layout."
        )
    return rows


def parse_addendum_a(xlsx_bytes: bytes, source_file: str) -> list[tuple]:
    """Parse Addendum A. Returns rows for opps_addendum_a."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_strs = [str(c).strip().upper() if c is not None else "" for c in row]
        if (
            "APC" in row_strs
            and any(s in ("GROUP TITLE", "TITLE", "DESCRIPTION") for s in row_strs)
            and "PAYMENT RATE" in row_strs
        ):
            header_row = row
            header_row_idx = i
            break

    if header_row is None:
        print("ERROR: Could not find Addendum A header row")
        return []

    print(f"  Addendum A header at row {header_row_idx}: {[str(c)[:20] for c in header_row[:8]]}")

    apc_col = find_col(header_row, "APC")
    title_col = find_col(header_row, "GROUP TITLE", "TITLE", "DESCRIPTION")
    si_col = find_col(header_row, "STATUS INDICATOR", "SI")
    rw_col = find_col(header_row, "RELATIVE WEIGHT")
    rate_col = find_col(header_row, "PAYMENT RATE", "PAYMENT", "RATE")
    cop_col = find_col(header_row, "NATIONAL UNADJ", "NATIONAL COPAY")
    min_cop_col = find_col(header_row, "MINIMUM UNADJ", "MIN COPAY")
    ira_col = find_col(header_row, "IRA", "COINSURANCE")
    adj_cop_col = find_col(header_row, "ADJUSTED BENEF", "ADJ COPAY")
    exp_col = find_col(header_row, "EXPIRATION")
    note_col = find_col(header_row, "NOTE")

    rows: list[tuple] = []
    data_started = False

    for row in ws.iter_rows(values_only=True):
        if not data_started:
            # Detect data start by APC column position instead of exact header row equality.
            # This keeps parsing stable if CMS adds or renames unrelated columns.
            if apc_col is not None and apc_col < len(row) and row[apc_col] is not None:
                candidate = str(row[apc_col]).strip()
                if re.match(r"^[0-9]{4}$", candidate):
                    data_started = True
                    # Fall through — process this row as data.
                else:
                    continue
            else:
                continue

        if apc_col is None or apc_col >= len(row):
            continue

        apc = str(row[apc_col]).strip() if row[apc_col] is not None else ""
        if not apc or not re.match(r"^[0-9]{4}$", apc):
            continue

        def get(col) -> str | float | None:
            if col is None or col >= len(row):
                return None
            return row[col]

        rows.append((
            QUARTER, EFFECTIVE_DATE, apc,
            str(get(title_col) or "").strip() or None,
            str(get(si_col) or "").strip() or None,
            safe_float(get(rw_col)),
            safe_float(get(rate_col)),
            safe_float(get(cop_col)),
            safe_float(get(min_cop_col)),
            safe_float(get(ira_col)),
            safe_float(get(adj_cop_col)),
            str(get(exp_col) or "").strip() or None,
            str(get(note_col) or "").strip() or None,
            source_file,
        ))

    wb.close()
    return rows


def get_xlsx_from_zip(zip_bytes: bytes) -> tuple[bytes, str]:
    """Extract the first .xlsx from a ZIP. Returns (xlsx_bytes, filename)."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        print(f"  ZIP contents: {archive.namelist()}")
        xlsx_files = [n for n in archive.namelist() if n.lower().endswith(".xlsx")
                      and not n.startswith("__")]
        if not xlsx_files:
            raise RuntimeError("No .xlsx file found in ZIP")
        fname = xlsx_files[0]
        return archive.read(fname), fname


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--addb", help="Local Addendum B ZIP path")
    parser.add_argument("--adda", help="Local Addendum A ZIP path")
    args = parser.parse_args()

    os.makedirs(DB_PATH.parent, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    create_schema(conn)

    # --- Addendum B ---
    print("\nProcessing Addendum B (HCPCS -> APC -> payment)...")
    if args.addb:
        zip_b = Path(args.addb).read_bytes()
    else:
        print(f"Downloading {ADDB_URL} ...")
        zip_b = download_bytes(ADDB_URL)
        print(f"  Downloaded {len(zip_b):,} bytes")

    xlsx_b, fname_b = get_xlsx_from_zip(zip_b)
    print(f"Parsing {fname_b} ...")
    rows_b = parse_addendum_b(xlsx_b, fname_b)
    print(f"Parsed {len(rows_b):,} Addendum B rows")

    inserted_b = 0
    for row in rows_b:
        try:
            conn.execute("""INSERT OR REPLACE INTO opps_addendum_b
                (quarter, effective_date, hcpcs_code, short_descriptor, status_indicator, apc,
                 relative_weight, payment_rate, national_unadjusted_copayment,
                 minimum_unadjusted_copayment, ira_coinsurance_percentage,
                 adjusted_beneficiary_copayment, pass_through_expiration, note, changed_flag, source_file)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", row)
            inserted_b += 1
        except sqlite3.Error as e:
            print(f"  Insert B error: {e}")

    conn.commit()
    print(f"Inserted {inserted_b:,} Addendum B rows")

    # --- Addendum A ---
    print("\nProcessing Addendum A (APC reference)...")
    if args.adda:
        zip_a = Path(args.adda).read_bytes()
    else:
        print(f"Downloading {ADDA_URL} ...")
        zip_a = download_bytes(ADDA_URL)
        print(f"  Downloaded {len(zip_a):,} bytes")

    xlsx_a, fname_a = get_xlsx_from_zip(zip_a)
    print(f"Parsing {fname_a} ...")
    rows_a = parse_addendum_a(xlsx_a, fname_a)
    print(f"Parsed {len(rows_a):,} Addendum A rows")

    inserted_a = 0
    for row in rows_a:
        try:
            conn.execute("""INSERT OR REPLACE INTO opps_addendum_a
                (quarter, effective_date, apc, group_title, status_indicator,
                 relative_weight, payment_rate, national_unadjusted_copayment,
                 minimum_unadjusted_copayment, ira_coinsurance_percentage,
                 adjusted_beneficiary_copayment, pass_through_expiration, note, source_file)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", row)
            inserted_a += 1
        except sqlite3.Error as e:
            print(f"  Insert A error: {e}")

    conn.commit()
    print(f"Inserted {inserted_a:,} Addendum A rows")

    # Summary
    print(f"\n{'='*50}")
    b_count = conn.execute("SELECT COUNT(*) FROM opps_addendum_b").fetchone()[0]
    a_count = conn.execute("SELECT COUNT(*) FROM opps_addendum_a").fetchone()[0]
    print(f"opps_addendum_b: {b_count:,} rows")
    print(f"opps_addendum_a: {a_count:,} rows")

    # Spot checks
    row = conn.execute(
        "SELECT hcpcs_code, short_descriptor, status_indicator, apc, payment_rate FROM opps_addendum_b WHERE hcpcs_code='99285'"
    ).fetchone()
    print(f"\n99285 in Addendum B: {row}")

    row = conn.execute(
        "SELECT hcpcs_code, short_descriptor, status_indicator, apc, payment_rate FROM opps_addendum_b WHERE hcpcs_code='70450'"
    ).fetchone()
    print(f"70450 in Addendum B: {row}")

    conn.close()
    size_kb = DB_PATH.stat().st_size // 1024
    print(f"\nWrote {DB_PATH} ({size_kb:,} KB)")


if __name__ == "__main__":
    main()
