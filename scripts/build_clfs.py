#!/usr/bin/env python3
"""
Build CLFS (Clinical Laboratory Fee Schedule) lookup JSON.
Downloads the CMS CLFS release ZIP and extracts HCPCS/CPT code -> payment rate.

Lab codes paid under CLFS are used as a fallback when MPFS has no rate.
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
import sys
import urllib.request
import zipfile
from pathlib import Path

OUTPUT_PATH = Path(__file__).parent.parent / "src" / "lib" / "data" / "clfs.json"

CLFS_URLS = [
    "https://www.cms.gov/files/zip/26clabq2.zip",
    "https://www.cms.gov/files/zip/26clabq1.zip",
]

CODE_PATTERN = re.compile(r"^(?:[0-9]{5}|[0-9]{4}[A-Z]|[A-Z][0-9]{4})$")


def parse_clfs_delimited(raw_bytes: bytes, delimiter: str) -> dict[str, dict[str, object]]:
    """Parse CLFS CSV/TXT content and return code -> { rate, description }."""
    text = io.StringIO(raw_bytes.decode("utf-8-sig", errors="replace"))
    reader = csv.reader(text, delimiter=delimiter)

    header = None
    for row in reader:
        normalized = [str(cell).strip().upper() if cell is not None else "" for cell in row]
        if "HCPCS" in normalized and "RATE" in normalized:
            header = normalized
            break

    if not header:
        print("ERROR: Could not find CLFS header row.")
        return {}

    def find_index(*names: str) -> int | None:
        for idx, name in enumerate(header):
            if name in names:
                return idx
        return None

    code_idx = find_index("HCPCS")
    rate_idx = find_index("RATE", "PAYMENT RATE")
    short_desc_idx = find_index("SHORTDESC", "SHORT DESCRIPTOR", "SHORT DESCRIPTION")
    long_desc_idx = find_index("LONGDESC", "LONG DESCRIPTION", "LONG DESCRIPTOR")
    extended_desc_idx = find_index("EXTENDEDLONGDESC", "EXTENDED LONG DESCRIPTION")

    if code_idx is None or rate_idx is None:
        print("ERROR: Missing HCPCS or RATE columns in CLFS file.")
        return {}

    rates: dict[str, dict[str, object]] = {}
    for row in reader:
        if code_idx >= len(row) or rate_idx >= len(row):
            continue

        code = str(row[code_idx]).strip().upper()
        if not CODE_PATTERN.match(code):
            continue

        raw_rate = str(row[rate_idx]).strip().replace(",", "")
        if not raw_rate:
            continue

        try:
            rate = float(raw_rate)
        except ValueError:
            continue

        if rate <= 0:
            continue

        description = ""
        for idx in (extended_desc_idx, long_desc_idx, short_desc_idx):
            if idx is not None and idx < len(row):
                candidate = str(row[idx]).strip()
                if candidate:
                    description = candidate
                    break

        rates[code] = {
            "rate": round(rate, 2),
            "description": description,
        }

    return rates


def parse_clfs_xlsx(xlsx_bytes: bytes) -> dict[str, dict[str, object]]:
    """Fallback XLSX parser for CLFS releases."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return {}

    rates: dict[str, dict[str, object]] = {}
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    header = None
    for row in ws.iter_rows(values_only=True):
        normalized = [str(cell).strip().upper() if cell is not None else "" for cell in row]
        if "HCPCS" in normalized and "RATE" in normalized:
            header = normalized
            break

    if not header:
        print("ERROR: Could not find HCPCS/RATE header row in XLSX.")
        return {}

    code_idx = header.index("HCPCS")
    rate_idx = header.index("RATE")
    short_desc_idx = header.index("SHORTDESC") if "SHORTDESC" in header else None
    long_desc_idx = header.index("LONGDESC") if "LONGDESC" in header else None
    extended_desc_idx = header.index("EXTENDEDLONGDESC") if "EXTENDEDLONGDESC" in header else None

    for row in ws.iter_rows(values_only=True):
        if not row or code_idx >= len(row) or row[code_idx] is None:
            continue

        code = str(row[code_idx]).strip().upper()
        if not CODE_PATTERN.match(code):
            continue

        raw_rate = row[rate_idx] if rate_idx < len(row) else None
        try:
            rate = float(raw_rate)
        except (TypeError, ValueError):
            continue

        if rate <= 0:
            continue

        description = ""
        for idx in (extended_desc_idx, long_desc_idx, short_desc_idx):
            if idx is not None and idx < len(row) and row[idx]:
                description = str(row[idx]).strip()
                break

        rates[code] = {
            "rate": round(rate, 2),
            "description": description,
        }

    return rates


def download_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        return resp.read()


def main() -> None:
    os.makedirs(OUTPUT_PATH.parent, exist_ok=True)

    local_file = sys.argv[1] if len(sys.argv) > 1 else None
    if local_file and Path(local_file).exists():
        print(f"Using local file: {local_file}")
        zip_bytes = Path(local_file).read_bytes()
        source_used = str(local_file)
    else:
        zip_bytes = None
        source_used = ""
        for url in CLFS_URLS:
            print(f"Trying {url}...")
            try:
                zip_bytes = download_bytes(url)
                source_used = url
                print(f"Downloaded {len(zip_bytes):,} bytes")
                break
            except Exception as exc:
                print(f"  Failed: {exc}")

    if not zip_bytes:
        print("ERROR: All downloads failed.")
        print(
            "Manual download page: "
            "https://www.cms.gov/medicare/payment/fee-schedules/clinical-laboratory-fee-schedule-clfs/files"
        )
        print("Then run: python3 scripts/build_clfs.py /path/to/downloaded.zip")
        sys.exit(1)

    rates: dict[str, dict[str, object]] = {}
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        print(f"Files in ZIP: {archive.namelist()}")

        for filename in archive.namelist():
            lower = filename.lower()
            if lower.endswith(".csv"):
                print(f"Parsing {filename}...")
                rates = parse_clfs_delimited(archive.read(filename), ",")
            elif lower.endswith(".txt"):
                print(f"Parsing {filename}...")
                rates = parse_clfs_delimited(archive.read(filename), "~")
            elif lower.endswith(".xlsx"):
                print(f"Parsing {filename}...")
                rates = parse_clfs_xlsx(archive.read(filename))
            else:
                continue

            if rates:
                print(f"Parsed {len(rates):,} rates from {filename}")
                break

    if not rates:
        print("ERROR: No rates parsed. Check if the CLFS format changed.")
        sys.exit(1)

    OUTPUT_PATH.write_text(json.dumps(rates, sort_keys=True, indent=2))
    size_kb = OUTPUT_PATH.stat().st_size // 1024
    print(f"Wrote {len(rates):,} rates to {OUTPUT_PATH} ({size_kb} KB)")
    print(f"Source used: {source_used}")
    for code in ["85025", "80053", "36415", "0001U", "85610"]:
        print(f"  {code}: {rates.get(code)}")


if __name__ == "__main__":
    main()
