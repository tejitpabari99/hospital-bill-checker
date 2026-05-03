#!/usr/bin/env python3
"""
Build data/dmepos.sqlite from CMS DMEPOS fee schedule.

Source (Q2 2026 / April): https://www.cms.gov/files/zip/dme26-b.zip
Inner file: DMEPOS_APR.xlsx

Two tables:
  dmepos_base       — one row per HCPCS/mod/mod2 combination
  dmepos_state_rates — non-rural state rates (one row per state per base row)

Stage 1: non-rural state rates only.

Usage:
  python3 scripts/build_dmepos_sqlite.py
  python3 scripts/build_dmepos_sqlite.py /path/to/dme26-b.zip
"""

from __future__ import annotations

import io
import os
import re
import sqlite3
import sys
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    sys.exit(1)

DB_PATH = Path(__file__).parent.parent / "data" / "dmepos.sqlite"

DMEPOS_URL = "https://www.cms.gov/files/zip/dme26-b.zip"
QUARTER = "2026Q2"
EFFECTIVE_DATE = "2026-04-01"
SOURCE_URL = "https://www.cms.gov/medicare/payment/fee-schedules/dmepos/dmepos-fee-schedule"

CODE_PATTERN = re.compile(r"^(?:[0-9]{5}|[0-9]{4}[A-Z]|[A-Z][0-9]{4})$")

# All valid US state/territory codes
VALID_STATES = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
    "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
    "VA","WA","WV","WI","WY","DC","PR","VI","GU","AS","MP",
}


def download_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        return resp.read()


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS dmepos_base (
            dmepos_id       INTEGER PRIMARY KEY AUTOINCREMENT,
            quarter         TEXT NOT NULL,
            effective_date  TEXT NOT NULL,
            hcpcs_code      TEXT NOT NULL,
            mod             TEXT,
            mod2            TEXT,
            jurisdiction    TEXT,
            category        TEXT,
            ceiling         NUMERIC,
            floor           NUMERIC,
            description     TEXT,
            source_file     TEXT,
            source_url      TEXT,
            updated_at      TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_dmepos_base_code
            ON dmepos_base(hcpcs_code);

        CREATE INDEX IF NOT EXISTS idx_dmepos_base_code_mods
            ON dmepos_base(hcpcs_code, mod, mod2);

        CREATE TABLE IF NOT EXISTS dmepos_state_rates (
            dmepos_id       INTEGER NOT NULL,
            state_code      TEXT NOT NULL,
            fee_amount      NUMERIC,
            PRIMARY KEY (dmepos_id, state_code),
            FOREIGN KEY (dmepos_id) REFERENCES dmepos_base(dmepos_id)
        );

        CREATE INDEX IF NOT EXISTS idx_dmepos_state_lookup
            ON dmepos_state_rates(dmepos_id, state_code);

        CREATE INDEX IF NOT EXISTS idx_dmepos_state_code
            ON dmepos_state_rates(state_code);
    """)


def wipe_existing_rows(conn: sqlite3.Connection) -> None:
    # Wipe existing rows so re-runs are idempotent.
    # dmepos_state_rates must be deleted first due to the FK constraint on dmepos_base.hcpcs_code.
    conn.execute("DELETE FROM dmepos_state_rates")
    conn.execute("DELETE FROM dmepos_base")
    conn.commit()


def safe_float(val) -> float | None:
    if val is None:
        return None
    s = str(val).replace(",", "").replace("$", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def find_col(header: list[str], *candidates: str) -> int | None:
    normalized = [str(h).strip().upper() if h is not None else "" for h in header]
    for candidate in candidates:
        cand_upper = candidate.upper()
        for i, h in enumerate(normalized):
            if h == cand_upper:
                return i
    return None


def parse_state_columns(header: list[str]) -> dict[str, int]:
    """Return {state_code: col_index} for all STATE (NR) columns."""
    state_cols: dict[str, int] = {}
    for i, h in enumerate(header):
        h_str = str(h).strip().upper() if h is not None else ""
        match = re.match(r"^([A-Z]{2})\s*\(NR\)$", h_str)
        if match:
            state = match.group(1)
            if state in VALID_STATES:
                state_cols[state] = i
    return state_cols


def parse_dmepos_xlsx(xlsx_bytes: bytes, source_file: str) -> tuple[list[tuple], dict[int, dict[str, float | None]]]:
    """
    Parse DMEPOS_APR.xlsx.
    Returns:
      base_rows: list of (quarter, eff_date, hcpcs, mod, mod2, juris, catg, ceil, floor, desc, source, url, updated_at)
      state_rates: {row_index: {state_code: fee_amount}}
    """
    updated_at = datetime.now(timezone.utc).isoformat()
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_strs = [str(c).strip().upper() if c is not None else "" for c in row]
        if "HCPCS" in row_strs:
            header_row = list(row)
            header_row_idx = i
            break

    if header_row is None:
        raise RuntimeError("Could not find DMEPOS header row")

    print(f"  Header at row {header_row_idx}: {[str(c)[:15] for c in header_row[:12]]}")

    hcpcs_col = find_col(header_row, "HCPCS")
    mod_col = find_col(header_row, "MOD", "MODIFIER")
    mod2_col = find_col(header_row, "MOD2", "MODIFIER2", "MOD 2")
    juris_col = find_col(header_row, "JURIS", "JURISDICTION")
    catg_col = find_col(header_row, "CATG", "CATEGORY")
    ceil_col = find_col(header_row, "CEILING")
    floor_col = find_col(header_row, "FLOOR")
    desc_col = find_col(header_row, "DESCRIPTION", "DESC")

    state_cols = parse_state_columns(header_row)
    print(f"  Found {len(state_cols)} state (NR) columns: {sorted(state_cols.keys())[:5]}...")

    base_rows: list[tuple] = []
    state_data: dict[int, dict[str, float | None]] = {}
    row_idx = 0
    data_started = False

    for row in ws.iter_rows(values_only=True):
        if not data_started:
            if tuple(row) == tuple(header_row):
                data_started = True
            continue

        if hcpcs_col is None or hcpcs_col >= len(row) or row[hcpcs_col] is None:
            continue

        code = str(row[hcpcs_col]).strip().upper()
        if not CODE_PATTERN.match(code):
            continue

        def get_str(col) -> str | None:
            if col is None or col >= len(row) or row[col] is None:
                return None
            s = str(row[col]).strip()
            return s if s else None

        mod = get_str(mod_col)
        mod2 = get_str(mod2_col)
        juris = get_str(juris_col)
        catg = get_str(catg_col)
        ceiling = safe_float(row[ceil_col] if ceil_col is not None and ceil_col < len(row) else None)
        floor_val = safe_float(row[floor_col] if floor_col is not None and floor_col < len(row) else None)
        desc = get_str(desc_col)

        base_rows.append((
            QUARTER, EFFECTIVE_DATE, code, mod, mod2, juris, catg,
            ceiling, floor_val, desc, source_file, SOURCE_URL, updated_at
        ))

        row_states: dict[str, float | None] = {}
        for state, col in state_cols.items():
            if col < len(row):
                fee = safe_float(row[col])
                if fee is not None:
                    row_states[state] = fee

        state_data[row_idx] = row_states
        row_idx += 1

    wb.close()
    return base_rows, state_data


def main() -> None:
    os.makedirs(DB_PATH.parent, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    create_schema(conn)

    if len(sys.argv) > 1:
        zip_bytes = Path(sys.argv[1]).read_bytes()
        print(f"Using local file: {sys.argv[1]}")
    else:
        print(f"Downloading {DMEPOS_URL} ...")
        zip_bytes = download_bytes(DMEPOS_URL)
        print(f"  Downloaded {len(zip_bytes):,} bytes")

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        print(f"ZIP contents: {archive.namelist()}")
        xlsx_files = [n for n in archive.namelist() if n.lower().endswith(".xlsx") and "DMEPOS" in n.upper()]
        if not xlsx_files:
            xlsx_files = [n for n in archive.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx_files:
            print("ERROR: No XLSX in ZIP")
            sys.exit(1)

        fname = xlsx_files[0]
        print(f"Parsing {fname} ...")
        base_rows, state_data = parse_dmepos_xlsx(archive.read(fname), fname)
    print(f"Parsed {len(base_rows):,} DMEPOS base rows")

    wipe_existing_rows(conn)

    inserted_base = 0
    dmepos_ids: list[int] = []
    for row in base_rows:
        try:
            cursor = conn.execute("""INSERT INTO dmepos_base
                (quarter, effective_date, hcpcs_code, mod, mod2, jurisdiction, category,
                 ceiling, floor, description, source_file, source_url, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""", row)
            dmepos_ids.append(cursor.lastrowid)
            inserted_base += 1
        except sqlite3.Error as e:
            dmepos_ids.append(-1)
            print(f"  Base insert error: {e}")

    conn.commit()
    print(f"Inserted {inserted_base:,} base rows")

    inserted_states = 0
    for row_idx, (dmepos_id, states) in enumerate(zip(dmepos_ids, [state_data.get(i, {}) for i in range(len(dmepos_ids))])):
        if dmepos_id < 0:
            continue
        for state, fee in states.items():
            if fee is None:
                continue
            try:
                conn.execute("""INSERT OR REPLACE INTO dmepos_state_rates
                    (dmepos_id, state_code, fee_amount) VALUES (?,?,?)""",
                    (dmepos_id, state, fee))
                inserted_states += 1
            except sqlite3.Error:
                pass

    conn.commit()
    print(f"Inserted {inserted_states:,} state rate rows")

    base_count = conn.execute("SELECT COUNT(*) FROM dmepos_base").fetchone()[0]
    state_count = conn.execute("SELECT COUNT(*) FROM dmepos_state_rates").fetchone()[0]
    print(f"\ndmepos_base: {base_count:,}")
    print(f"dmepos_state_rates: {state_count:,}")

    row = conn.execute("""
        SELECT b.hcpcs_code, b.description, r.state_code, r.fee_amount
        FROM dmepos_base b
        JOIN dmepos_state_rates r ON r.dmepos_id = b.dmepos_id
        WHERE b.hcpcs_code = 'E0601' AND r.state_code = 'TX'
        LIMIT 1
    """).fetchone()
    print(f"\nE0601 (CPAP) in TX: {row}")

    conn.close()
    size_kb = DB_PATH.stat().st_size // 1024
    print(f"\nWrote {DB_PATH} ({size_kb:,} KB)")


if __name__ == "__main__":
    main()
