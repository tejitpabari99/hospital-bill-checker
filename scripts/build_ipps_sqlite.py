#!/usr/bin/env python3
"""
Build data/ipps.sqlite from CMS IPPS FY2026 Table 5 MS-DRG weights.

Source: https://www.cms.gov/files/zip/fy2026-ipps-fr-table-5.zip

Usage:
  python3 scripts/build_ipps_sqlite.py
  python3 scripts/build_ipps_sqlite.py /path/to/fy2026-ipps-fr-table-5.zip
"""

from __future__ import annotations

import io
import os
import re
import sqlite3
import sys
import urllib.request
import zipfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    sys.exit(1)

DB_PATH = Path(__file__).parent.parent / "data" / "ipps.sqlite"

IPPS_URL = "https://www.cms.gov/files/zip/fy2026-ipps-fr-table-5.zip"
FISCAL_YEAR = "2026"
SOURCE_URL = "https://www.cms.gov/medicare/payment/prospective-payment-systems/acute-inpatient-pps/fy-2026-ipps-final-rule-home-page"

DRG_PATTERN = re.compile(r"^\d{1,3}$")


def download_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        return resp.read()


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS ipps_drg_rates (
            fiscal_year         TEXT NOT NULL,
            ms_drg              TEXT NOT NULL,
            title               TEXT,
            relative_weight     NUMERIC,
            geometric_mean_los  NUMERIC,
            arithmetic_mean_los NUMERIC,
            source_file         TEXT,
            source_url          TEXT,
            updated_at          TEXT NOT NULL,
            PRIMARY KEY (fiscal_year, ms_drg)
        );

        CREATE INDEX IF NOT EXISTS idx_ipps_drg_code
            ON ipps_drg_rates(ms_drg);

        CREATE TABLE IF NOT EXISTS ipps_payment_parameters (
            fiscal_year                 TEXT PRIMARY KEY,
            operating_standardized_amount NUMERIC,
            capital_base_rate           NUMERIC,
            source_file                 TEXT,
            source_url                  TEXT,
            updated_at                  TEXT NOT NULL
        );
    """)


def find_col(header_row, *candidates: str) -> int | None:
    """Case-insensitive partial match."""
    for i, cell in enumerate(header_row):
        cell_upper = str(cell).strip().upper() if cell is not None else ""
        for candidate in candidates:
            if candidate.upper() in cell_upper:
                return i
    return None


def parse_table5(xlsx_bytes: bytes, source_file: str) -> list[tuple]:
    """
    Parse Table 5 Excel.
    Find header row containing 'MS-DRG', then parse data.
    Returns list of (fiscal_year, ms_drg, title, rel_weight, gmlos, amlos, source_file, source_url, updated_at)
    """
    from datetime import datetime, timezone

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    # Try each sheet - CMS sometimes puts data on a non-active sheet
    header_row = None
    ws_used = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_strs = [str(c).strip().upper() if c is not None else "" for c in row]
            # Look for the actual column header row, not the CMS title row.
            has_drg = any("MS-DRG" in s or (s == "DRG" and i < 10) for s in row_strs)
            has_title = find_col(row, "TITLE", "DESCRIPTION", "MS-DRG TITLE") is not None
            has_weight_or_los = (
                find_col(row, "WEIGHT", "RELATIVE WEIGHT", "WEIGHTS") is not None
                or find_col(row, "GEOMETRIC", "GMLOS", "GEO MEAN") is not None
            )
            if has_drg and has_title and has_weight_or_los:
                header_row = row
                ws_used = ws
                print(f"  Header found on sheet '{sheet_name}' at row {i}: {[str(c)[:20] for c in row[:8]]}")
                break
        if header_row is not None:
            break

    if header_row is None or ws_used is None:
        print("ERROR: Could not find DRG header row in any sheet")
        return []

    drg_col = find_col(header_row, "MS-DRG", "DRG")
    title_col = find_col(header_row, "TITLE", "DESCRIPTION", "MS-DRG TITLE")
    rw_col = find_col(header_row, "WEIGHT", "RELATIVE WEIGHT", "WEIGHTS")
    gmlos_col = find_col(header_row, "GEOMETRIC", "GMLOS", "GEO MEAN")
    amlos_col = find_col(header_row, "ARITHMETIC", "AMLOS", "ARITH MEAN")

    print(f"  Columns - DRG:{drg_col}, Title:{title_col}, Weight:{rw_col}, GMLOS:{gmlos_col}, AMLOS:{amlos_col}")

    updated_at = datetime.now(timezone.utc).isoformat()
    rows: list[tuple] = []
    data_started = False

    for row in ws_used.iter_rows(values_only=True):
        if not data_started:
            if row == tuple(header_row):
                data_started = True
            continue

        if drg_col is None or drg_col >= len(row) or row[drg_col] is None:
            continue

        drg_raw = str(row[drg_col]).strip()
        if not DRG_PATTERN.match(drg_raw):
            continue

        # Zero-pad to 3 digits
        ms_drg = drg_raw.zfill(3)

        def safe_float(val) -> float | None:
            if val is None:
                return None
            try:
                return float(str(val).replace(",", "").strip())
            except (ValueError, TypeError):
                return None

        title = str(row[title_col]).strip() if title_col is not None and title_col < len(row) and row[title_col] else None
        rel_weight = safe_float(row[rw_col] if rw_col is not None and rw_col < len(row) else None)
        gmlos = safe_float(row[gmlos_col] if gmlos_col is not None and gmlos_col < len(row) else None)
        amlos = safe_float(row[amlos_col] if amlos_col is not None and amlos_col < len(row) else None)

        if rel_weight is None:
            continue

        rows.append((
            FISCAL_YEAR, ms_drg, title, rel_weight, gmlos, amlos,
            source_file, SOURCE_URL, updated_at
        ))

    wb.close()
    return rows


def main() -> None:
    os.makedirs(DB_PATH.parent, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    create_schema(conn)

    if len(sys.argv) > 1:
        zip_bytes = Path(sys.argv[1]).read_bytes()
        print(f"Using local file: {sys.argv[1]}")
    else:
        print(f"Downloading {IPPS_URL} ...")
        zip_bytes = download_bytes(IPPS_URL)
        print(f"  Downloaded {len(zip_bytes):,} bytes")

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        print(f"ZIP contents: {archive.namelist()}")
        xlsx_files = [n for n in archive.namelist() if n.lower().endswith(".xlsx")
                      and not n.startswith("__")]
        if not xlsx_files:
            print("ERROR: No .xlsx in ZIP")
            sys.exit(1)

        fname = xlsx_files[0]
        print(f"Parsing {fname} ...")
        rows = parse_table5(archive.read(fname), fname)
        print(f"Parsed {len(rows):,} DRG rows")

    conn.execute("DELETE FROM ipps_drg_rates WHERE fiscal_year = ?", (FISCAL_YEAR,))

    # Insert
    inserted = 0
    for row in rows:
        try:
            conn.execute("""INSERT OR REPLACE INTO ipps_drg_rates
                (fiscal_year, ms_drg, title, relative_weight, geometric_mean_los,
                 arithmetic_mean_los, source_file, source_url, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?)""", row)
            inserted += 1
        except sqlite3.Error as e:
            print(f"  Insert error: {e}")

    conn.commit()
    print(f"\nInserted {inserted:,} MS-DRG rows")

    # Verify
    for drg in ["470", "291", "001", "470"]:
        row = conn.execute(
            "SELECT ms_drg, title, relative_weight, geometric_mean_los FROM ipps_drg_rates WHERE ms_drg=? AND fiscal_year=?",
            (drg, FISCAL_YEAR)
        ).fetchone()
        print(f"DRG {drg}: {row}")

    conn.close()
    size_kb = DB_PATH.stat().st_size // 1024
    print(f"\nWrote {DB_PATH} ({size_kb:,} KB)")
    print("\nNote: ipps_payment_parameters table created but not populated (use for full payment calc in future).")


if __name__ == "__main__":
    main()
