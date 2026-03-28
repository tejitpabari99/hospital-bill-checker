#!/usr/bin/env python3
"""
Build NCCI (National Correct Coding Initiative) PTP lookup JSON.
Downloads quarterly NCCI PTP edits from CMS.

NCCI PTP table columns:
  Column 1 Code: comprehensive code (gets paid)
  Column 2 Code: component code (bundled into Column 1, should NOT be billed separately)
  Modifier Indicator: 0=cannot override, 1=modifier -59 can override

Output format: { "column2_code": "column1_code" }
So we can check: if billed_code in ncci → it's bundled into ncci[billed_code]
"""
import json
import zipfile
import io
import os
import re
import urllib.request
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

OUTPUT_PATH = Path(__file__).parent.parent / "src" / "lib" / "data" / "ncci.json"

# CMS NCCI PTP files — check https://www.cms.gov/medicare/coding-billing/national-correct-coding-initiative-ncci-edits for latest
# File format: XLSX inside ZIP
NCCI_URL = "https://www.cms.gov/medicare/coding-billing/national-correct-coding-initiative-ncci-edits/downloads/ncci-ptp-edits-current.zip"

def parse_ncci_xlsx(xlsx_bytes: bytes) -> dict[str, str]:
    """Parse NCCI XLSX and return {col2_code: col1_code} mapping."""
    if not HAS_OPENPYXL:
        print("openpyxl not installed — install with: pip install openpyxl")
        return {}

    bundling: dict[str, str] = {}
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # Find header row
        header_row_idx = None
        col1_idx = col2_idx = mod_idx = None

        for i, row in enumerate(rows[:10]):
            row_text = ' '.join(str(c).upper() for c in row if c)
            if 'COLUMN 1' in row_text or 'COL 1' in row_text or 'COLUMN1' in row_text:
                header_row_idx = i
                for j, cell in enumerate(row):
                    if cell:
                        ct = str(cell).upper()
                        if 'COLUMN 1' in ct or 'COL1' in ct: col1_idx = j
                        elif 'COLUMN 2' in ct or 'COL2' in ct: col2_idx = j
                        elif 'MODIFIER' in ct: mod_idx = j
                break

        if header_row_idx is None or col1_idx is None or col2_idx is None:
            continue

        for row in rows[header_row_idx + 1:]:
            if not row or not row[col1_idx] or not row[col2_idx]:
                continue
            col1 = str(row[col1_idx]).strip()
            col2 = str(row[col2_idx]).strip()

            # Validate CPT format
            if re.match(r'^[0-9]{5}$|^[JGABC][0-9]{4}$', col1) and re.match(r'^[0-9]{5}$|^[JGABC][0-9]{4}$', col2):
                bundling[col2] = col1

        if bundling:
            print(f"Parsed {len(bundling)} NCCI pairs from sheet '{sheet_name}'")
            break

    return bundling

def main():
    os.makedirs(OUTPUT_PATH.parent, exist_ok=True)

    print(f"Downloading NCCI from {NCCI_URL}...")
    try:
        req = urllib.request.Request(NCCI_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=60) as resp:
            zip_bytes = resp.read()
        print(f"Downloaded {len(zip_bytes):,} bytes")

        bundling: dict[str, str] = {}
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            xlsx_files = [n for n in z.namelist() if n.lower().endswith('.xlsx')]
            print(f"XLSX files in ZIP: {xlsx_files}")

            for xlsx_name in xlsx_files:
                with z.open(xlsx_name) as f:
                    xlsx_bytes = f.read()
                pairs = parse_ncci_xlsx(xlsx_bytes)
                bundling.update(pairs)
                if bundling:
                    break

    except Exception as e:
        print(f"Download failed: {e}")
        print("Generating minimal fallback with common NCCI pairs...")
        # Common NCCI bundling pairs for fallback
        bundling = {
            "27370": "27447",  # knee injection bundled into total knee replacement
            "93010": "93000",  # ECG interpretation bundled into complete ECG
            "36000": "36410",  # IV access bundled into venipuncture
            "99070": "99213",  # supplies bundled into E&M
        }

    OUTPUT_PATH.write_text(json.dumps(bundling, indent=2))
    print(f"Wrote {len(bundling):,} NCCI pairs to {OUTPUT_PATH}")

if __name__ == '__main__':
    main()
