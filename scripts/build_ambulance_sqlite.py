#!/usr/bin/env python3
"""
Build data/ambulance.sqlite from CMS Ambulance Fee Schedule PUF and geography file.

Sources:
  Rate file: https://www.cms.gov/files/zip/cy-2026-file.zip
    Inner: Copy_of_AFS2026_PUF_ext.xlsx
  Geography: https://www.cms.gov/files/zip/zip-code-carrier-locality-file-revised-02-18-2026.zip
    Inner: Geographic_Area_2026.xlsx

Usage:
  python3 scripts/build_ambulance_sqlite.py
  python3 scripts/build_ambulance_sqlite.py --rates /path/to/cy-2026-file.zip --geo /path/to/geo.zip
"""

from __future__ import annotations

import argparse
import io
import os
import re
import sqlite3
import sys
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    sys.exit(1)

DB_PATH = Path(__file__).parent.parent / "data" / "ambulance.sqlite"

RATE_URL = "https://www.cms.gov/files/zip/cy-2026-file.zip"
GEO_URL = "https://www.cms.gov/files/zip/zip-code-carrier-locality-file-revised-02-18-2026.zip"

CALENDAR_YEAR = "2026"
SOURCE_RATE_URL = "https://www.cms.gov/medicare/payment/fee-schedules/ambulance/ambulance-fee-schedule-public-use-files"
SOURCE_GEO_URL = "https://www.cms.gov/medicare/payment/fee-schedules"

# Ambulance HCPCS codes start with A0 for transport services
AMBULANCE_CODE_PATTERN = re.compile(r"^A0[0-9]{3}$")
GENERAL_CODE_PATTERN = re.compile(r"^(?:[0-9]{5}|[0-9]{4}[A-Z]|[A-Z][0-9]{4})$")


def download_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        return resp.read()


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS ambulance_rates (
            ambulance_rate_id   INTEGER PRIMARY KEY AUTOINCREMENT,
            calendar_year       TEXT NOT NULL,
            hcpcs_code          TEXT NOT NULL,
            short_description   TEXT,
            locality            TEXT,
            area_type           TEXT,
            base_rate           NUMERIC,
            mileage_rate        NUMERIC,
            rate_amount         NUMERIC,
            source_file         TEXT,
            source_url          TEXT,
            updated_at          TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_ambulance_rates_code
            ON ambulance_rates(hcpcs_code);

        CREATE INDEX IF NOT EXISTS idx_ambulance_rates_code_locality
            ON ambulance_rates(hcpcs_code, locality);

        CREATE TABLE IF NOT EXISTS ambulance_geography (
            zip_code    TEXT PRIMARY KEY,
            locality    TEXT,
            state       TEXT,
            area_type   TEXT,
            source_file TEXT,
            updated_at  TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_ambulance_geo_locality
            ON ambulance_geography(locality);

        CREATE INDEX IF NOT EXISTS idx_ambulance_geo_state
            ON ambulance_geography(state);
    """)


def find_col(header_row, *candidates: str) -> int | None:
    for i, cell in enumerate(header_row):
        cell_upper = str(cell).strip().upper() if cell is not None else ""
        for candidate in candidates:
            if candidate.upper() in cell_upper:
                return i
    return None


def safe_float(val) -> float | None:
    if val is None:
        return None
    s = str(val).replace(",", "").replace("$", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_rates_xlsx(xlsx_bytes: bytes, source_file: str) -> list[tuple]:
    """
    Parse AFS2026_PUF_ext.xlsx.
    We don't know the exact column names until we inspect, so we do flexible header detection.
    Returns list of (year, hcpcs, description, locality, area_type, base_rate, mileage_rate, rate_amount, file, url, updated_at)
    """
    updated_at = datetime.now(timezone.utc).isoformat()
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    rows: list[tuple] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find header row
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_strs = [str(c).strip().upper() if c is not None else "" for c in row]
            if "HCPCS" in row_strs or any("HCPCS" in s for s in row_strs):
                header_row = list(row)
                print(f"  Rates header on sheet '{sheet_name}' row {i}: {[str(c)[:20] for c in header_row[:10]]}")
                break

        if header_row is None:
            print(f"  Skipping sheet '{sheet_name}' - no HCPCS column")
            continue

        hcpcs_col = find_col(header_row, "HCPCS")
        desc_col = find_col(header_row, "DESCRIPTION", "DESC", "SERVICE")
        locality_col = find_col(header_row, "LOCALITY", "AREA", "GEOGRAPHIC")
        area_type_col = find_col(header_row, "AREA TYPE", "URBAN", "RURAL", "SUPER RURAL")
        base_col = find_col(header_row, "BASE RATE", "BASE", "GROUND BASE")
        mileage_col = find_col(header_row, "MILEAGE", "MILE")
        rate_col = find_col(header_row, "RATE", "PAYMENT", "AMOUNT")
        # If both base and mileage exist, use those; otherwise fall back to rate_col

        data_started = False
        for row in ws.iter_rows(values_only=True):
            if not data_started:
                if tuple(row) == tuple(header_row):
                    data_started = True
                continue

            if hcpcs_col is None or hcpcs_col >= len(row) or row[hcpcs_col] is None:
                continue

            code = str(row[hcpcs_col]).strip().upper()
            if not GENERAL_CODE_PATTERN.match(code):
                continue

            def get_str(col) -> str | None:
                if col is None or col >= len(row) or row[col] is None:
                    return None
                s = str(row[col]).strip()
                return s if s else None

            description = get_str(desc_col)
            locality = get_str(locality_col)
            area_type = get_str(area_type_col)
            base_rate = safe_float(row[base_col] if base_col is not None and base_col < len(row) else None)
            mileage_rate = safe_float(row[mileage_col] if mileage_col is not None and mileage_col < len(row) else None)
            rate_amount = safe_float(row[rate_col] if rate_col is not None and rate_col < len(row) else None)

            # If we have both base and mileage, the generic rate_amount may be redundant
            if base_rate is None and mileage_rate is None and rate_amount is None:
                continue

            rows.append((
                CALENDAR_YEAR, code, description, locality, area_type,
                base_rate, mileage_rate, rate_amount,
                source_file, SOURCE_RATE_URL, updated_at
            ))

        break  # Use first sheet with data

    wb.close()
    return rows


def parse_geography_xlsx(xlsx_bytes: bytes, source_file: str) -> list[tuple]:
    """
    Parse Geographic_Area_2026.xlsx.
    Returns list of (zip_code, locality, state, area_type, source_file, updated_at)
    """
    updated_at = datetime.now(timezone.utc).isoformat()
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    rows: list[tuple] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_strs = [str(c).strip().upper() if c is not None else "" for c in row]
            if "ZIP" in row_strs or any("ZIP" in s for s in row_strs):
                header_row = list(row)
                print(f"  Geo header on sheet '{sheet_name}' row {i}: {[str(c)[:20] for c in header_row[:10]]}")
                break

        if header_row is None:
            print(f"  Skipping sheet '{sheet_name}' - no ZIP column")
            continue

        zip_col = find_col(header_row, "ZIP", "ZIP CODE", "ZIPCODE")
        locality_col = find_col(header_row, "LOCALITY", "AREA", "GEOGRAPHIC AREA")
        state_col = find_col(header_row, "STATE")
        area_type_col = find_col(header_row, "AREA TYPE", "URBAN", "RURAL")

        data_started = False
        for row in ws.iter_rows(values_only=True):
            if not data_started:
                if tuple(row) == tuple(header_row):
                    data_started = True
                continue

            if zip_col is None or zip_col >= len(row) or row[zip_col] is None:
                continue

            zip_raw = str(row[zip_col]).strip().zfill(5)[:5]
            if not re.match(r"^\d{5}$", zip_raw):
                continue

            def get_str(col) -> str | None:
                if col is None or col >= len(row) or row[col] is None:
                    return None
                s = str(row[col]).strip()
                return s if s else None

            locality = get_str(locality_col)
            state = get_str(state_col)
            area_type = get_str(area_type_col)

            rows.append((zip_raw, locality, state, area_type, source_file, updated_at))

        break  # Use first sheet with data

    wb.close()
    return rows


def get_xlsx_from_zip(zip_bytes: bytes, name_hint: str = "") -> tuple[bytes, str]:
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        print(f"  ZIP contents: {archive.namelist()}")
        xlsx_files = [
            n for n in archive.namelist()
            if n.lower().endswith(".xlsx") and not n.startswith("__")
        ]
        if name_hint:
            preferred = [n for n in xlsx_files if name_hint.lower() in n.lower()]
            if preferred:
                xlsx_files = preferred

        if not xlsx_files:
            raise RuntimeError("No .xlsx file found in ZIP")
        fname = xlsx_files[0]
        return archive.read(fname), fname


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rates", help="Local rates ZIP path")
    parser.add_argument("--geo", help="Local geography ZIP path")
    args = parser.parse_args()

    os.makedirs(DB_PATH.parent, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    create_schema(conn)

    # --- Rates ---
    print("\nProcessing ambulance rates...")
    if args.rates:
        rates_zip = Path(args.rates).read_bytes()
    else:
        print(f"Downloading {RATE_URL} ...")
        rates_zip = download_bytes(RATE_URL)
        print(f"  Downloaded {len(rates_zip):,} bytes")

    xlsx_bytes, fname = get_xlsx_from_zip(rates_zip, "AFS")
    print(f"Parsing {fname} ...")
    rate_rows = parse_rates_xlsx(xlsx_bytes, fname)
    print(f"Parsed {len(rate_rows):,} rate rows")

    inserted_rates = 0
    for row in rate_rows:
        try:
            conn.execute("""INSERT INTO ambulance_rates
                (calendar_year, hcpcs_code, short_description, locality, area_type,
                 base_rate, mileage_rate, rate_amount, source_file, source_url, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""", row)
            inserted_rates += 1
        except sqlite3.Error as e:
            print(f"  Rate insert error: {e}")

    conn.commit()
    print(f"Inserted {inserted_rates:,} rate rows")

    # --- Geography ---
    print("\nProcessing ambulance geography...")
    if args.geo:
        geo_zip = Path(args.geo).read_bytes()
    else:
        print(f"Downloading {GEO_URL} ...")
        geo_zip = download_bytes(GEO_URL)
        print(f"  Downloaded {len(geo_zip):,} bytes")

    geo_xlsx, geo_fname = get_xlsx_from_zip(geo_zip, "Geographic")
    print(f"Parsing {geo_fname} ...")
    geo_rows = parse_geography_xlsx(geo_xlsx, geo_fname)
    print(f"Parsed {len(geo_rows):,} geography rows")

    inserted_geo = 0
    for row in geo_rows:
        try:
            conn.execute("""INSERT OR REPLACE INTO ambulance_geography
                (zip_code, locality, state, area_type, source_file, updated_at)
                VALUES (?,?,?,?,?,?)""", row)
            inserted_geo += 1
        except sqlite3.Error as e:
            print(f"  Geo insert error: {e}")

    conn.commit()
    print(f"Inserted {inserted_geo:,} geography rows")

    # Summary
    print(f"\n{'='*50}")
    print(f"ambulance_rates: {conn.execute('SELECT COUNT(*) FROM ambulance_rates').fetchone()[0]:,}")
    print(f"ambulance_geography: {conn.execute('SELECT COUNT(*) FROM ambulance_geography').fetchone()[0]:,}")

    # Spot check
    row = conn.execute("SELECT * FROM ambulance_geography WHERE zip_code='90210' LIMIT 1").fetchone()
    print(f"\nZIP 90210: {row}")
    row = conn.execute("SELECT hcpcs_code, locality, base_rate, rate_amount FROM ambulance_rates WHERE hcpcs_code='A0427' LIMIT 3").fetchall()
    print(f"A0427 (ALS1) rates: {row}")

    conn.close()
    size_kb = DB_PATH.stat().st_size // 1024
    print(f"\nWrote {DB_PATH} ({size_kb:,} KB)")


if __name__ == "__main__":
    main()
